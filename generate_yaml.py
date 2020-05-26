#!/usr/bin/env python3
from openpyxl import load_workbook
import sys

LEVEL_NUMBERS = ['0.1', '0.2', '1', '2', '3', '4', '5']

def main():
    wb = load_workbook('input.xlsx')
    for level_number, level_name in zip(LEVEL_NUMBERS, wb.sheetnames[:len(LEVEL_NUMBERS)]):
        ws = wb[level_name]
        for word_column in 'AB':
            if ws[word_column + '1'].value == '詞彙':
                break
        
        row = 2
        while True:
            word = ws[word_column + str(row)].value
            if not word:
                break
            print(level_number + ',' + word)
            row += 1


if __name__ == '__main__' and not hasattr(sys, 'ps1'):
    main()
